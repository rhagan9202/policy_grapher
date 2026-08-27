"""Read the reference CSV into an immutable, database-free structure.

Everything here is pure: no Neo4j, no settings, no I/O beyond reading the file.
"""

import ast
import csv
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from policy_grapher.sources import SourceError

EXPECTED_COLUMNS = ["Document Name", "References", "Type"]

_PARENTHESISED = re.compile(r"\([^)]*\)")
_NON_ALPHANUMERIC = re.compile(r"[^a-z0-9]")


class CsvSourceError(SourceError):
    """The CSV could not be parsed as expected."""


def canonical_name(name: str) -> str:
    """Aggressive normalisation used only to spot near-duplicate names."""
    without_parens = _PARENTHESISED.sub("", name)
    return _NON_ALPHANUMERIC.sub("", without_parens.casefold())


@dataclass(frozen=True)
class CorpusRow:
    name: str
    reference_role: str
    references: tuple[str, ...]


@dataclass(frozen=True)
class ParsedCorpus:
    rows: tuple[CorpusRow, ...]
    corpus_names: frozenset[str]
    external_names: frozenset[str]
    edges: tuple[tuple[str, str], ...]
    self_references_skipped: int
    suspected_duplicates: tuple[tuple[str, ...], ...]

    @property
    def all_names(self) -> frozenset[str]:
        return self.corpus_names | self.external_names


def _parse_references(cell: str, document: str) -> tuple[str, ...]:
    try:
        value = ast.literal_eval(cell)
    except (ValueError, SyntaxError) as exc:
        raise CsvSourceError(
            f"References for {document!r} are not a valid Python list: {exc}"
        ) from exc
    if not isinstance(value, (list, tuple)):
        raise CsvSourceError(
            f"References for {document!r} parsed as {type(value).__name__}, not a list."
        )
    return tuple(str(item).strip() for item in value if str(item).strip())


def _find_suspected_duplicates(names: frozenset[str]) -> tuple[tuple[str, ...], ...]:
    groups: dict[str, list[str]] = defaultdict(list)
    for name in sorted(names):
        groups[canonical_name(name)].append(name)
    return tuple(
        tuple(group) for _, group in sorted(groups.items()) if len(group) > 1
    )


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != EXPECTED_COLUMNS:
            raise CsvSourceError(
                f"Expected columns {EXPECTED_COLUMNS}, found {reader.fieldnames}."
            )
        return list(reader)


def _read_xlsx_rows(path: Path) -> list[dict[str, str]]:
    """The same three columns, out of a spreadsheet — STORY-036.

    Imported here rather than at module scope so that reading a CSV does not
    depend on a spreadsheet library being importable, and so a broken openpyxl
    fails the format that needs it instead of every format.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError, zipfile.BadZipFile) as exc:
        # A renamed file, a corrupt one, or a `.xls` in disguise. The library's
        # own message names none of them usefully, and a traceback is not an
        # answer for someone who renamed something.
        raise CsvSourceError(f"Could not read {path.name} as a spreadsheet.") from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise CsvSourceError(f"{path.name} has no rows at all, not even a header.")

    found = [None if cell is None else str(cell).strip() for cell in header]
    # Trailing empty cells are ordinary in a spreadsheet and mean nothing.
    while found and found[-1] in (None, ""):
        found.pop()
    if found != EXPECTED_COLUMNS:
        raise CsvSourceError(f"Expected columns {EXPECTED_COLUMNS}, found {found}.")

    return [
        {
            column: "" if value is None else str(value)
            for column, value in zip(EXPECTED_COLUMNS, row, strict=False)
        }
        for row in rows
        # A spreadsheet keeps blank rows a CSV would simply not have.
        if any(cell is not None and str(cell).strip() for cell in row)
    ]


READERS = {".csv": _read_csv_rows, ".xlsx": _read_xlsx_rows}


def parse_corpus(path: Path) -> ParsedCorpus:
    """A manifest, from whichever format it arrived in.

    Both readers return the same rows, and everything after this line is shared —
    so the two formats cannot describe different corpora by drifting apart, only
    by disagreeing about how to read a file. `test_manifest_xlsx.py` asserts they
    do not.
    """
    reader = READERS.get(path.suffix.lower())
    if reader is None:
        raise CsvSourceError(
            f"{path.name} is not a manifest this reads: expected one of "
            f"{sorted(READERS)}."
        )
    raw_rows = reader(path)

    rows: list[CorpusRow] = []
    edges: set[tuple[str, str]] = set()
    self_references = 0

    for raw in raw_rows:
        name = (raw["Document Name"] or "").strip()
        if not name:
            raise CsvSourceError("A row has an empty 'Document Name'.")
        role = (raw["Type"] or "").strip()
        references = _parse_references(raw["References"] or "[]", name)

        rows.append(CorpusRow(name=name, reference_role=role, references=references))

        for target in references:
            if target == name:
                self_references += 1
                continue
            edges.add((name, target))

    corpus_names = frozenset(row.name for row in rows)
    referenced = frozenset(t for _, t in edges) | frozenset(
        target for row in rows for target in row.references
    )
    external_names = referenced - corpus_names

    return ParsedCorpus(
        rows=tuple(rows),
        corpus_names=corpus_names,
        external_names=external_names,
        edges=tuple(sorted(edges)),
        self_references_skipped=self_references,
        suspected_duplicates=_find_suspected_duplicates(corpus_names | external_names),
    )
